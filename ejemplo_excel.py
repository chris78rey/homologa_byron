#!/usr/bin/env python3
"""Genera un Excel de ejemplo para homologación."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Homologación ISSFA"

# Encabezados
headers = ['CODIGO_ACTUAL', 'DESCRIPCION_ACTUAL', 'CODIGO_NUEVO', 'DESCRIPCION_NUEVA', 'TIPO']
for col, header in enumerate(headers, 1):
    cell = ws.cell(1, col, header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="366092")
    cell.alignment = Alignment(horizontal="center")

# Datos de ejemplo
datos = [
    ('C08CA01-01', 'AMLODIPINA - SÓLIDO ORAL - TABLETA - 5 MG', 'C08CA01-01-N', 'AMLODIPINA, TABLETA 5 MG', 'M'),
    ('C08CA01-02', 'AMLODIPINA - SÓLIDO ORAL - TABLETA - 10 MG', 'C08CA01-02-N', 'AMLODIPINA, TABLETA 10 MG', 'M'),
    ('D07AC01-01', 'BETAMETASONA - SEMISÓLIDO CUTÁNEO - CREMA - 0,05 % - 15 G', 'D07AC01-01-N', 'BETAMETASONA, CREMA 0,05% TUBO X 15G', 'M'),
    ('H05BX02-01', 'PARICALCITOL - LÍQUIDO PARENTERAL - SOLUCION INYECTABLE', 'H05BX02-01-N', 'PARICALCITOL, SOL INYECTABLE 5 MCG/ML', 'M'),
]

for row_idx, datos_fila in enumerate(datos, 2):
    for col, valor in enumerate(datos_fila, 1):
        ws.cell(row_idx, col, valor)

# Ajustar ancho de columnas
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 8

wb.save("ejemplo_homologacion.xlsx")
print("Excel de ejemplo creado: ejemplo_homologacion.xlsx")
