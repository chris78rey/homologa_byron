"""
Lógica de homologación y algoritmos de similitud de texto.
Formato Excel: CODIGO_ACTUAL, DESCRIPCION_ACTUAL, CODIGO_NUEVO, DESCRIPCION_NUEVA, TIPO
"""
import csv
from datetime import datetime
from dataclasses import dataclass, field
from typing import Optional
import jellyfish


@dataclass
class SimilitudResultado:
    """Resultado del análisis de similitud."""
    score_jaro_winkler: float = 0.0
    score_levenshtein: float = 0.0
    confianza: str = "BAJA"
    recomendacion: str = "DESCARTAR"
    
    @classmethod
    def calcular(cls, texto1: str, texto2: str) -> 'SimilitudResultado':
        """Calcula similitud entre dos textos."""
        if not texto1 or not texto2:
            return cls(0.0, 0.0, "BAJA", "DESCARTAR")
        
        t1 = cls._normalizar(texto1)
        t2 = cls._normalizar(texto2)
        
        jw_score = jellyfish.jaro_winkler_similarity(t1, t2) * 100
        lev_score = cls._levenshtein_ratio(t1, t2) * 100
        
        score_promedio = (jw_score * 0.7) + (lev_score * 0.3)
        
        if score_promedio >= 97:
            confianza = "ALTA"
            recomendacion = "APLICAR"
        elif score_promedio >= 88:
            confianza = "MEDIA"
            recomendacion = "REVISAR"
        else:
            confianza = "BAJA"
            recomendacion = "DESCARTAR"
        
        return cls(jw_score, lev_score, confianza, recomendacion)
    
    @staticmethod
    def _normalizar(texto: str) -> str:
        return texto.upper().strip()
    
    @staticmethod
    def _levenshtein_ratio(s1: str, s2: str) -> float:
        if not s1 and not s2:
            return 1.0
        if not s1 or not s2:
            return 0.0
        max_len = max(len(s1), len(s2))
        distance = jellyfish.levenshtein_distance(s1, s2)
        return 1.0 - (distance / max_len)


@dataclass
class HomologacionItem:
    """Representa un item de homologación del Excel."""
    fila_excel: int
    codigo_actual: str
    descripcion_actual_excel: str
    codigo_nuevo: str
    descripcion_nueva: str
    tipo: str = "M"
    
    # Datos de Oracle
    existe_actual: bool = False
    existe_nuevo: bool = False
    descripcion_actual_oracle: str = ""
    tipo_oracle: str = ""
    
    # Similitudes
    score_oracle_excel: float = 0.0
    score_actual_nueva: float = 0.0
    
    # Resultado del análisis
    accion: str = ""
    status: str = ""
    aplicar: bool = False
    decision: str = ""
    
    # Equivalencias
    tiene_equivalencias: bool = False
    
    @classmethod
    def from_excel_row(cls, fila: int, row: dict) -> 'HomologacionItem':
        return cls(
            fila_excel=fila,
            codigo_actual=str(row.get('CODIGO_ACTUAL', '')).strip(),
            descripcion_actual_excel=str(row.get('DESCRIPCION_ACTUAL', '')).strip(),
            codigo_nuevo=str(row.get('CODIGO_NUEVO', '')).strip(),
            descripcion_nueva=str(row.get('DESCRIPCION_NUEVA', '')).strip(),
            tipo=str(row.get('TIPO', 'M')).strip() or 'M',
        )


class HomologacionEngine:
    """Motor de homologación de items ISSFA."""
    
    def __init__(self, db, id_itisf: int = 1, threshold: float = 88.0):
        self.db = db
        self.id_itisf = id_itisf
        self.threshold = threshold
        self.items: list[HomologacionItem] = []
        self.backup_date = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.backup_exitoso = False
    
    def load_excel(self, filepath: str) -> int:
        """Carga items desde archivo Excel."""
        import openpyxl
        
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        headers = [str(cell.value).strip().upper() if cell.value else "" for cell in ws[1]]
        
        # Mapear columnas con aliases
        col_map = self._map_columns(headers)
        
        self.items = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_name, col_idx in col_map.items():
                if col_idx < len(headers):
                    row_data[col_name] = ws.cell(row_idx, col_idx + 1).value or ""
            
            # Saltar filas vacías
            codigo_actual = row_data.get('CODIGO_ACTUAL', '')
            codigo_nuevo = row_data.get('CODIGO_NUEVO', '')
            desc_nueva = row_data.get('DESCRIPCION_NUEVA', '')
            
            if not codigo_nuevo or str(codigo_nuevo).lower() == 'nan':
                continue
            if not desc_nueva or str(desc_nueva).lower() == 'nan':
                continue
            
            item = HomologacionItem.from_excel_row(row_idx, row_data)
            self.items.append(item)
        
        return len(self.items)
    
    def _map_columns(self, headers: list) -> dict:
        """Mapea columnas del Excel a nombres canónicos."""
        aliases = {
            'CODIGO_ACTUAL': ['CODIGO_ACTUAL', 'COD_ACTUAL', 'CODIGO_ANTERIOR', 'CODIGO_ORIGINAL'],
            'DESCRIPCION_ACTUAL': ['DESCRIPCION_ACTUAL', 'DESC_ACTUAL', 'DESCRIPCION_ANTERIOR'],
            'CODIGO_NUEVO': ['CODIGO_NUEVO', 'COD_NUEVO', 'CODIGO_ISSFA', 'CODIGO_FINAL'],
            'DESCRIPCION_NUEVA': ['DESCRIPCION_NUEVA', 'DESC_NUEVA', 'DESCRIPCION_ISSFA'],
            'TIPO': ['TIPO', 'TIPO_ITEM', 'CLASE'],
        }
        
        upper_headers = {h: i for i, h in enumerate(headers)}
        result = {}
        
        for canonical, names in aliases.items():
            for name in names:
                if name in upper_headers:
                    result[canonical] = upper_headers[name]
                    break
        
        return result
    
    def _find_detalle(self, codigo: str) -> dict:
        """Busca código en ITEMS_ISSFA_DETALLE."""
        try:
            result = self.db.execute(
                """SELECT ID_ITISF, CODIGO, DESCRIPCION, TIPO
                   FROM SIS.ITEMS_ISSFA_DETALLE
                   WHERE ID_ITISF = :1 AND CODIGO = :2""",
                (self.id_itisf, codigo)
            )
            if result:
                return {
                    'ID_ITISF': result[0][0],
                    'CODIGO': result[0][1],
                    'DESCRIPCION': result[0][2],
                    'TIPO': result[0][3]
                }
        except Exception:
            pass
        return None
    
    def _check_equivalencias(self, codigo_issfa: str) -> bool:
        """Verifica si el código tiene equivalencias asociadas."""
        try:
            result = self.db.execute(
                """SELECT COUNT(*) FROM SIS.EQUIVALENCIAS_ITEMS_ISSFA
                   WHERE ID_ITISF = :1 AND CODIGO_ISSFA = :2""",
                (self.id_itisf, codigo_issfa)
            )
            return result[0][0] > 0 if result else False
        except Exception:
            return False
    
    def _decision_from_score(self, score: float) -> str:
        """Determina decisión basada en score."""
        if score >= 97:
            return "ALTA_CONFIANZA"
        elif score >= 88:
            return "REVISAR"
        else:
            return "NO_RECOMENDADO"
    
    def crear_backup(self) -> bool:
        """Crea backup de tablas."""
        try:
            bkp_detalle = f"SIS.BKP_ITEMS_ISSFA_DETALLE_{self.backup_date}"
            bkp_equiv = f"SIS.BKP_EQUIVALENCIAS_ITEMS_ISSFA_{self.backup_date}"
            
            try:
                self.db.execute(f"DROP TABLE {bkp_detalle}")
            except Exception:
                pass
            try:
                self.db.execute(f"DROP TABLE {bkp_equiv}")
            except Exception:
                pass
            
            self.db.execute(f"CREATE TABLE {bkp_detalle} AS SELECT * FROM SIS.ITEMS_ISSFA_DETALLE")
            self.db.execute(f"CREATE TABLE {bkp_equiv} AS SELECT * FROM SIS.EQUIVALENCIAS_ITEMS_ISSFA")
            self.db.commit()
            self.backup_exitoso = True
            return True
        except Exception as e:
            print(f"Error backup: {e}")
            return False
    
    def restaurar_backup(self) -> bool:
        """Restaura tablas desde backup."""
        if not self.backup_exitoso:
            return False
        try:
            bkp_detalle = f"SIS.BKP_ITEMS_ISSFA_DETALLE_{self.backup_date}"
            bkp_equiv = f"SIS.BKP_EQUIVALENCIAS_ITEMS_ISSFA_{self.backup_date}"
            
            self.db.execute("DELETE FROM SIS.EQUIVALENCIAS_ITEMS_ISSFA")
            self.db.execute(f"INSERT INTO SIS.EQUIVALENCIAS_ITEMS_ISSFA SELECT * FROM {bkp_equiv}")
            self.db.execute("DELETE FROM SIS.ITEMS_ISSFA_DETALLE")
            self.db.execute(f"INSERT INTO SIS.ITEMS_ISSFA_DETALLE SELECT * FROM {bkp_detalle}")
            self.db.commit()
            return True
        except Exception as e:
            print(f"Error restore: {e}")
            self.db.rollback_force()
            return False
    
    def analizar(self) -> dict:
        """Analiza todos los items y determina acciones."""
        stats = {
            'total': len(self.items),
            'update': 0,
            'insert': 0,
            'bloqueado': 0,
            'alta_confianza': 0,
            'media_confianza': 0,
            'baja_confianza': 0,
        }
        
        for item in self.items:
            # Buscar en Oracle
            actual_db = self._find_detalle(item.codigo_actual)
            nuevo_db = self._find_detalle(item.codigo_nuevo)
            
            item.existe_actual = actual_db is not None
            item.existe_nuevo = nuevo_db is not None
            
            if item.existe_actual:
                item.descripcion_actual_oracle = actual_db.get('DESCRIPCION', '')
                item.tipo_oracle = actual_db.get('TIPO', 'M')
                if not item.tipo:
                    item.tipo = item.tipo_oracle
                
                # Calcular similitud Oracle vs Excel
                item.score_oracle_excel = SimilitudResultado.calcular(
                    item.descripcion_actual_oracle,
                    item.descripcion_actual_excel
                ).score_jaro_winkler
            
            # Calcular similitud Actual vs Nueva
            item.score_actual_nueva = SimilitudResultado.calcular(
                item.descripcion_actual_excel,
                item.descripcion_nueva
            ).score_jaro_winkler
            
            score_max = max(item.score_oracle_excel, item.score_actual_nueva)
            item.decision = self._decision_from_score(score_max)
            
            # Determinar acción
            if item.existe_actual and not item.existe_nuevo:
                item.tiene_equivalencias = self._check_equivalencias(item.codigo_actual)
                
                if item.tiene_equivalencias:
                    item.accion = "BLOQUEADO"
                    item.status = "TIENE_EQUIVALENCIAS"
                    item.aplicar = False
                    stats['bloqueado'] += 1
                else:
                    item.accion = "UPDATE"
                    item.status = "ACTUALIZAR_CODIGO_Y_DESC"
                    item.aplicar = (item.score_oracle_excel >= self.threshold and 
                                   item.score_actual_nueva >= self.threshold)
                    stats['update'] += 1
                
            elif item.existe_actual and item.existe_nuevo and item.codigo_actual != item.codigo_nuevo:
                item.accion = "BLOQUEADO"
                item.status = "CODIGO_NUEVO_YA_EXISTE"
                item.aplicar = False
                stats['bloqueado'] += 1
                
            elif item.existe_actual and item.existe_nuevo and item.codigo_actual == item.codigo_nuevo:
                item.accion = "UPDATE"
                item.status = "ACTUALIZAR_SOLO_DESC"
                item.aplicar = item.score_oracle_excel >= self.threshold
                stats['update'] += 1
                
            elif not item.existe_actual and not item.existe_nuevo:
                if item.tipo:
                    item.accion = "INSERT"
                    item.status = "INSERTAR_NUEVO"
                    item.aplicar = item.score_actual_nueva >= self.threshold
                    stats['insert'] += 1
                else:
                    item.accion = "BLOQUEADO"
                    item.status = "FALTA_TIPO"
                    item.aplicar = False
                    stats['bloqueado'] += 1
                    
            elif not item.existe_actual and item.existe_nuevo:
                item.accion = "BLOQUEADO"
                item.status = "CODIGO_NUEVO_YA_EXISTE_SIN_ACTUAL"
                item.aplicar = False
                stats['bloqueado'] += 1
            
            # Contar confidencias
            if item.decision == "ALTA_CONFIANZA":
                stats['alta_confianza'] += 1
            elif item.decision == "REVISAR":
                stats['media_confianza'] += 1
            else:
                stats['baja_confianza'] += 1
        
        return stats
    
    def aplicar_cambios(self) -> dict:
        """Aplica los cambios seleccionados a la base de datos."""
        resultados = {
            'updates': 0,
            'inserts': 0,
            'errores': [],
        }
        
        items_seleccionados = [i for i in self.items if i.aplicar]
        
        try:
            for item in items_seleccionados:
                if item.accion == "UPDATE":
                    self.db.execute(
                        """UPDATE SIS.ITEMS_ISSFA_DETALLE
                           SET CODIGO = :1, DESCRIPCION = :2
                           WHERE ID_ITISF = :3 AND CODIGO = :4""",
                        (item.codigo_nuevo, item.descripcion_nueva,
                         self.id_itisf, item.codigo_actual)
                    )
                    resultados['updates'] += 1
                    
                elif item.accion == "INSERT":
                    self.db.execute(
                        """INSERT INTO SIS.ITEMS_ISSFA_DETALLE
                               (ID_ITISF, CODIGO, DESCRIPCION, TIPO)
                           VALUES (:1, :2, :3, :4)""",
                        (self.id_itisf, item.codigo_nuevo, 
                         item.descripcion_nueva, item.tipo)
                    )
                    resultados['inserts'] += 1
            
            self.db.commit()
            
        except Exception as e:
            self.db.rollback_force()
            resultados['errores'].append(str(e))
            # NO restaurar automáticamente - solo botón manual
        
        return resultados
    
    def generar_csv_auditoria(self, filepath: str):
        """Genera archivo CSV con auditoría de cambios."""
        fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=';')
            
            writer.writerow([
                'FECHA', 'FILA_EXCEL', 'ACCION', 'STATUS', 'DECISION',
                'CODIGO_ACTUAL', 'DESC_ACTUAL_ORACLE', 'DESC_ACTUAL_EXCEL',
                'CODIGO_NUEVO', 'DESC_NUEVA',
                'TIPO', 'SCORE_ORACLE_EXCEL', 'SCORE_ACTUAL_NUEVA',
                'EXISTE_ACTUAL', 'EXISTE_NUEVO', 'TIENE_EQUIV',
                'APLICAR'
            ])
            
            for item in self.items:
                writer.writerow([
                    fecha,
                    item.fila_excel,
                    item.accion,
                    item.status,
                    item.decision,
                    item.codigo_actual,
                    item.descripcion_actual_oracle,
                    item.descripcion_actual_excel,
                    item.codigo_nuevo,
                    item.descripcion_nueva,
                    item.tipo,
                    f"{item.score_oracle_excel:.2f}",
                    f"{item.score_actual_nueva:.2f}",
                    'S' if item.existe_actual else 'N',
                    'S' if item.existe_nuevo else 'N',
                    'S' if item.tiene_equivalencias else 'N',
                    'S' if item.aplicar else 'N',
                ])
