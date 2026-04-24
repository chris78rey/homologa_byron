#!/usr/bin/env python3
"""Crea la plantilla oficial de homologación ISSFA."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "resources" / "templates" / "plantilla_homologacion_items_issfa.xlsx"


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    
    # === HOJA 1: Plantilla ===
    ws = wb.active
    ws.title = "Plantilla"
    
    headers = ["CODIGO_ACTUAL", "DESCRIPCION_ACTUAL", "CODIGO_NUEVO", "DESCRIPCION_NUEVA"]
    ws.append(headers)
    
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    
    widths = {"A": 22, "B": 60, "C": 22, "D": 60}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A2"
    
    # Filas vacías para llenar
    for row in range(2, 502):
        for col in range(1, 5):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    
    # === HOJA 2: Ejemplo ===
    ex = wb.create_sheet("Ejemplo")
    ex.append(headers)
    ex.append(["010101", "CONSULTA MEDICA GENERAL", "010101-A", "CONSULTA MÉDICA GENERAL"])
    ex.append(["020202", "RADIOGRAFIA TORAX", "020202-A", "RADIOGRAFÍA DE TÓRAX"])
    ex.append(["C08CA01-01", "AMLODIPINA TABLETA 5 MG", "C08CA01-01-N", "AMLODIPINA, TABLETA 5 MG"])
    
    for col in range(1, len(headers) + 1):
        cell = ex.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    
    for col, width in widths.items():
        ex.column_dimensions[col].width = width
    
    ex.freeze_panes = "A2"
    
    # === HOJA 3: Instrucciones ===
    ins = wb.create_sheet("Instrucciones")
    instructions = [
        ["Campo", "Descripción"],
        ["CODIGO_ACTUAL", "Código que actualmente existe en Oracle."],
        ["DESCRIPCION_ACTUAL", "Descripción actual esperada del código existente."],
        ["CODIGO_NUEVO", "Nuevo código que reemplazará al código actual o será insertado."],
        ["DESCRIPCION_NUEVA", "Nueva descripción que se desea registrar."],
        ["", ""],
        ["REGLAS IMPORTANTES", ""],
        ["", "No cambiar los nombres de las columnas."],
        ["", "No eliminar la fila 1 de encabezados."],
        ["", "El sistema mostrará una vista previa antes de aplicar cambios."],
        ["", "Se recomienda usar la plantilla oficial para evitar errores."],
    ]
    
    for row in instructions:
        ins.append(row)
    
    ins.column_dimensions["A"].width = 28
    ins.column_dimensions["B"].width = 90
    
    for col in range(1, 3):
        cell = ins.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    
    wb.save(OUTPUT)
    print(f"✅ Plantilla creada en: {OUTPUT}")


if __name__ == "__main__":
    main()
